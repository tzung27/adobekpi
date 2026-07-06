import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook('FY26_Q3_續約總表_範例.xlsx', data_only=True)
print('SHEETS:', wb.sheetnames)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'SHEET={sname}|COLS={ws.max_column}|ROWS={ws.max_row}')
    r1 = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    r2 = [str(c.value).strip() if c.value is not None else '' for c in ws[2]]
    print('ROW1=' + '|'.join(r1[:30]))
    print('ROW2=' + '|'.join(r2[:30]))
